#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ARM Kampagne - GaLaBau Leadliste Generator
Konsolidiert Daten aus GaLaBau-Research-Reports, priorisiert und exportiert
für CRM-Import und als Anrufliste.

Zielgruppe: Garten- und Landschaftsbauunternehmen (GaLaBau)
Use-Case: Frostschäden an Wegen, Einfahrten, Terrassen, Parkplätzen → ARM Kaltasphalt
"""

import csv
import io
import re
import os
from pathlib import Path

# Optional: openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("HINWEIS: openpyxl nicht installiert - Excel-Export wird übersprungen.")
    print("Installieren mit: pip install openpyxl\n")

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR

# PLZ region config: which source files to parse per region
PLZ_REGIONS = {
    "PLZ9": {
        "claude": "Recherche GaLaBau PLZ 9 durch claude.md",
        "label": "PLZ 9 (Bayern/Thüringen)",
    },
    "PLZ8": {
        "claude": "Recherche GaLaBau PLZ 8 durch claude.md",
        "label": "PLZ 8 (Bayern Süd - Oberbayern/Schwaben)",
    },
    "PLZ7": {
        "claude": "Recherche GaLaBau PLZ 7 durch claude.md",
        "label": "PLZ 7 (Baden-Württemberg)",
    },
    "PLZ6": {
        "claude": "Recherche GaLaBau PLZ 6 durch claude.md",
        "label": "PLZ 6 (Hessen/Saarland/Rhein-Neckar)",
    },
    "PLZ5": {
        "claude": "Recherche GaLaBau PLZ 5 durch claude.md",
        "label": "PLZ 5 (Köln/Bonn/Aachen/Koblenz/Trier)",
    },
    "PLZ4": {
        "claude": "Recherche GaLaBau PLZ 4 durch claude.md",
        "label": "PLZ 4 (NRW - Ruhrgebiet/Münsterland/Niederrhein)",
    },
    "PLZ3": {
        "claude": "Recherche GaLaBau PLZ 3 durch claude.md",
        "label": "PLZ 3 (Hannover/Braunschweig/Kassel/Göttingen/Magdeburg)",
    },
    "PLZ2": {
        "claude": "Recherche GaLaBau PLZ 2 durch claude.md",
        "label": "PLZ 2 (Hamburg/Schleswig-Holstein/Niedersachsen Nord/Bremen)",
    },
    "PLZ1": {
        "claude": "Recherche GaLaBau PLZ 1 durch claude.md",
        "label": "PLZ 1 (Berlin/Brandenburg/Mecklenburg-Vorpommern)",
    },
    "PLZ0": {
        "claude": "Recherche GaLaBau PLZ 0 durch claude.md",
        "label": "PLZ 0 (Sachsen/Sachsen-Anhalt/Thüringen)",
    },
}


# ---------------------------------------------------------------------------
# 1. DATA EXTRACTION
# ---------------------------------------------------------------------------

def parse_claude_markdown_tables(filepath):
    """Parse markdown tables from the Claude GaLaBau research report."""
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    records = []
    current_section = ""
    in_table = False
    headers = []

    for i, line in enumerate(lines):
        stripped = line.strip()

        # Detect section headers (GaLaBau-specific)
        if stripped.startswith("## 1."):
            current_section = "Kommune"
        elif stripped.startswith("## 2."):
            current_section = "Behörde"
        elif stripped.startswith("### Überregionale"):
            current_section = "Privat (Großunternehmen)"
        elif stripped.startswith("### Regionale"):
            current_section = "Privat (Mittelstand)"
        elif stripped.startswith("## 4."):
            current_section = "Landschaftsarchitekt"
        elif stripped.startswith("## Regionale Verteilung"):
            current_section = ""  # end of data

        # Detect table rows
        if stripped.startswith("|") and "---" not in stripped:
            cells = [c.strip().strip("*") for c in stripped.split("|")[1:-1]]

            if not in_table:
                # This is the header row
                headers = cells
                in_table = True
                continue

            if in_table and len(cells) >= 3:
                record = _parse_claude_table_row(cells, headers, current_section)
                if record:
                    records.append(record)
        elif in_table and not stripped.startswith("|"):
            in_table = False
            headers = []

    return records


def detect_bundesland(plz):
    """Detect Bundesland from PLZ prefix."""
    if not plz:
        return ""
    p = plz.strip()
    # Thüringen
    if p.startswith("98") or p.startswith("99"):
        return "Thüringen"
    if p.startswith("07") or p.startswith("046"):
        return "Thüringen"
    # Sachsen
    if p.startswith("01") or p.startswith("02") or p.startswith("04") or \
       p.startswith("08") or p.startswith("09"):
        return "Sachsen"
    # Sachsen-Anhalt
    if p.startswith("06"):
        return "Sachsen-Anhalt"
    # Bayern (PLZ 8 + PLZ 9)
    if p.startswith("8") or p.startswith("9"):
        return "Bayern"
    # Baden-Württemberg (PLZ 7)
    if p.startswith("7"):
        return "Baden-Württemberg"
    # Hessen
    if p.startswith("60") or p.startswith("61") or p.startswith("63") or \
       p.startswith("64") or p.startswith("65") or p.startswith("35") or \
       p.startswith("34") or p.startswith("36"):
        return "Hessen"
    # Saarland (PLZ 66)
    if p.startswith("66"):
        return "Saarland"
    # Rheinland-Pfalz
    if p.startswith("67") or p.startswith("55") or p.startswith("56") or \
       p.startswith("54") or p.startswith("53"):
        return "Rheinland-Pfalz"
    # Baden-Württemberg (PLZ 68, 69)
    if p.startswith("68") or p.startswith("69"):
        return "Baden-Württemberg"
    # NRW (PLZ 4 + PLZ 5)
    if p.startswith("4") or p.startswith("5"):
        return "Nordrhein-Westfalen"
    # Mecklenburg-Vorpommern (PLZ 17, 18, 19)
    if p.startswith("17") or p.startswith("18") or p.startswith("19"):
        return "Mecklenburg-Vorpommern"
    # Berlin (PLZ 10-13)
    if p.startswith("10") or p.startswith("12") or p.startswith("13"):
        return "Berlin"
    # Brandenburg (PLZ 14-16)
    if p.startswith("14") or p.startswith("15") or p.startswith("16"):
        return "Brandenburg"
    # Hamburg (PLZ 20-22)
    if p.startswith("20") or p.startswith("21") or p.startswith("22"):
        return "Hamburg"
    # Schleswig-Holstein (PLZ 23-25)
    if p.startswith("23") or p.startswith("24") or p.startswith("25"):
        return "Schleswig-Holstein"
    # Bremen (PLZ 28)
    if p.startswith("28"):
        return "Bremen"
    # Sachsen-Anhalt (PLZ 39)
    if p.startswith("39"):
        return "Sachsen-Anhalt"
    # Niedersachsen (PLZ 26, 27, 29, 3)
    if p.startswith("26") or p.startswith("27") or p.startswith("29") or p.startswith("3"):
        return "Niedersachsen"
    return ""


def _parse_claude_table_row(cells, headers, section):
    """Convert a row from Claude's GaLaBau markdown tables into our unified format."""
    record = {
        "source": "Claude",
        "kategorie": section,
        "name": "",
        "strasse": "",
        "plz": "",
        "ort": "",
        "bundesland": "",
        "zustaendigkeit": "",
        "leistungen": "",
        "kontaktperson": "",
        "telefon": "",
        "email": "",
    }

    if section == "Kommune":
        # Headers: Ort | PLZ | Bezeichnung | Ansprechpartner | Telefon | E-Mail / Website
        if len(cells) >= 6:
            record["ort"] = cells[0]
            record["plz"] = cells[1]
            record["name"] = cells[2]
            record["kontaktperson"] = cells[3] if cells[3] != "–" else ""
            record["telefon"] = cells[4]
            email_web = cells[5]
            emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', email_web)
            record["email"] = emails[0] if emails else ""
            record["leistungen"] = "Kommunale Grünflächenpflege, Wegebau, Flächeninstandsetzung"

    elif section == "Behörde":
        # Headers: Behörde/Einrichtung | PLZ / Ort | Zuständigkeitsbereich | Telefon | E-Mail / Website
        if len(cells) >= 5:
            record["name"] = cells[0]
            plz_ort = cells[1]
            plz_match = re.match(r'(\d{5})\s+(.*)', plz_ort)
            if plz_match:
                record["plz"] = plz_match.group(1)
                record["ort"] = plz_match.group(2)
            record["zustaendigkeit"] = cells[2]
            record["telefon"] = cells[3]
            email_web = cells[4]
            emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', email_web)
            record["email"] = emails[0] if emails else ""
            record["leistungen"] = "Öffentliche Grünanlagen, Parks, Gärten, Wegebau"

    elif section.startswith("Privat"):
        if "Großunternehmen" in section:
            # Headers: Unternehmen | PLZ / Ort | Leistungsspektrum | GF / Inhaber | Kontakt
            if len(cells) >= 5:
                record["name"] = cells[0]
                plz_ort = cells[1]
                plz_match = re.match(r'(\d{5})\s+(.*)', plz_ort)
                if plz_match:
                    record["plz"] = plz_match.group(1)
                    record["ort"] = plz_match.group(2)
                else:
                    record["ort"] = plz_ort
                record["leistungen"] = cells[2]
                record["kontaktperson"] = cells[3] if cells[3] != "–" else ""
                kontakt = cells[4]
                phones = re.findall(r'[\d\s/\-·]+(?:\d{2,})', kontakt)
                if phones:
                    record["telefon"] = phones[0].strip().strip("·").strip()
                emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', kontakt)
                record["email"] = emails[0] if emails else ""
        else:
            # Mittelstand: Unternehmen | PLZ / Ort | Leistungsspektrum | Kontakt
            if len(cells) >= 4:
                record["name"] = cells[0]
                plz_ort = cells[1]
                plz_match = re.match(r'(\d{5})\s+(.*)', plz_ort)
                if plz_match:
                    record["plz"] = plz_match.group(1)
                    record["ort"] = plz_match.group(2)
                else:
                    record["ort"] = plz_ort
                record["leistungen"] = cells[2]
                kontakt = cells[3]
                phones = re.findall(r'[\d\s/\-]+(?:\d{2,})', kontakt)
                if phones:
                    record["telefon"] = phones[0].strip()
                emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', kontakt)
                record["email"] = emails[0] if emails else ""
        record["kategorie"] = "Privat (GaLaBau)"

    elif section == "Landschaftsarchitekt":
        # Headers: Büro | PLZ / Ort | Leistungsspektrum | Leitung / Inhaber | Kontakt
        if len(cells) >= 5:
            record["name"] = cells[0]
            plz_ort = cells[1]
            plz_match = re.match(r'(\d{5})\s+(.*)', plz_ort)
            if plz_match:
                record["plz"] = plz_match.group(1)
                record["ort"] = plz_match.group(2)
            else:
                record["ort"] = plz_ort
            record["leistungen"] = cells[2]
            record["kontaktperson"] = cells[3] if cells[3] != "–" else ""
            kontakt = cells[4]
            phones = re.findall(r'[\d\s/\-]+(?:\d{2,})', kontakt)
            if phones:
                record["telefon"] = phones[0].strip()
            emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', kontakt)
            record["email"] = emails[0] if emails else ""

    # Skip empty records
    if not record["name"]:
        return None

    # Clean markdown artifacts from all fields
    for key in record:
        if isinstance(record[key], str):
            record[key] = record[key].replace("**", "").strip()

    # Auto-detect Bundesland from PLZ if not already set
    if not record["bundesland"] and record["plz"]:
        record["bundesland"] = detect_bundesland(record["plz"])

    return record


# ---------------------------------------------------------------------------
# 2. DEDUPLICATION
# ---------------------------------------------------------------------------

def normalize_name(name):
    """Normalize company name for dedup matching."""
    n = name.lower()
    for suffix in ["gmbh & co. kg", "gmbh & co.kg", "gmbh", "se", "ag",
                    "e.k.", "gbr", "ohg", "kg"]:
        n = n.replace(suffix, "")
    n = re.sub(r'[^a-zäöüß0-9\s]', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def extract_city_key(rec):
    """Extract a clean city name for city-based dedup of municipal orgs."""
    ort = rec.get("ort", "").strip()
    ort = re.sub(r'\*+', '', ort)
    ort = re.sub(r'\s*\(.*?\)', '', ort)
    ort = ort.split("/")[0].strip()
    name = rec.get("name", "")
    city_from_name = re.search(r'Stadt\s+(\w+)', name)
    if city_from_name:
        return city_from_name.group(1).lower().strip()
    return ort.lower().strip()


def _merge_records(winner, donor):
    """Merge contact data from donor into winner where winner has gaps."""
    for field in ["kontaktperson", "telefon", "email", "strasse",
                  "plz", "ort", "zustaendigkeit", "leistungen"]:
        if not winner[field] and donor[field]:
            winner[field] = donor[field]
    if winner["plz"] and "x" in winner["plz"].lower() and donor["plz"] and "x" not in donor["plz"].lower():
        winner["plz"] = donor["plz"]


def _contact_score(rec):
    """Score how much contact info a record has."""
    return sum([
        bool(rec["kontaktperson"] and rec["kontaktperson"] != "–"),
        bool(rec["telefon"]),
        bool(rec["email"]),
        bool(rec["strasse"]),
        bool(rec["plz"] and "x" not in rec["plz"].lower()),
    ])


def deduplicate(records):
    """Deduplicate records using multi-strategy matching."""
    # --- Strategy 1: City-based dedup for kommunale Grünflächenämter ---
    kommune_by_city = {}
    other_records = []

    for rec in records:
        kat = rec["kategorie"].lower()
        if "kommune" in kat:
            city_key = extract_city_key(rec)
            if not city_key:
                other_records.append(rec)
                continue
            if city_key in kommune_by_city:
                existing = kommune_by_city[city_key]
                if _contact_score(rec) > _contact_score(existing):
                    _merge_records(rec, existing)
                    kommune_by_city[city_key] = rec
                else:
                    _merge_records(existing, rec)
            else:
                kommune_by_city[city_key] = rec
        else:
            other_records.append(rec)

    # --- Strategy 2: Name-based dedup for non-kommune records ---
    seen = {}
    for rec in other_records:
        key = normalize_name(rec["name"])
        plz_prefix = rec["plz"][:3] if rec["plz"] and "x" not in rec["plz"][:3].lower() else ""
        dedup_key = f"{key}_{plz_prefix}"

        name_only_key = key
        if dedup_key in seen:
            existing = seen[dedup_key]
            if _contact_score(rec) > _contact_score(existing):
                _merge_records(rec, existing)
                seen[dedup_key] = rec
            else:
                _merge_records(existing, rec)
        elif name_only_key in seen:
            existing = seen[name_only_key]
            if _contact_score(rec) > _contact_score(existing):
                _merge_records(rec, existing)
                seen[name_only_key] = rec
            else:
                _merge_records(existing, rec)
        else:
            seen[dedup_key] = rec
            seen[name_only_key] = rec

    unique_others = list({id(v): v for v in seen.values()}.values())
    result = list(kommune_by_city.values()) + unique_others
    return result


# ---------------------------------------------------------------------------
# 3. PRIORITIZATION (GaLaBau-angepasst)
# ---------------------------------------------------------------------------

def categorize_priority(rec):
    """Assign A/B/C priority based on relevance for ARM GaLaBau campaign.

    A-Lead: Kommunales Grünflächenamt mit Telefon + E-Mail;
            Kleiner/mittlerer GaLaBau-Betrieb mit Fokus Wegebau/Pflaster + Telefon + E-Mail;
            high_relevance_names
    B-Lead: Größere GaLaBau-Ketten; Staatliche Gartenverwaltungen;
            Kommunen ohne vollständige Kontaktdaten
    C-Lead: Landschaftsarchitekten (indirekt); Einträge ohne Kontaktdaten
    """
    kategorie = rec["kategorie"].lower()
    name = rec["name"].lower()
    leistungen = rec["leistungen"].lower()
    has_contact = bool(rec["kontaktperson"] and rec["kontaktperson"] not in ["–", "-", ""])
    has_phone = bool(rec["telefon"])
    has_email = bool(rec["email"] and "@" in rec["email"])

    # GaLaBau-specific keywords indicating high relevance
    galabau_keywords = ["wegebau", "pflaster", "terrasse", "einfahrt", "parkplatz",
                        "asphalt", "instandsetzung", "sanierung", "reparatur",
                        "frostschad", "winterschad", "oberflächensanierung",
                        "flächeninstandsetzung", "wegesanierung", "belagarbeiten"]
    has_galabau_relevance = any(kw in leistungen for kw in galabau_keywords)

    # GaLaBau specialists known for path/surface repair work
    high_relevance_names = [
        # PLZ 9 (Bayern/Thüringen)
        "hauth galabau", "meysel", "brunner garten", "gries landschaft",
        "badum garten", "friedel garten", "zapf galabau", "böllert",
        "loibl garten", "kremsreiter", "knauer garten", "vicari",
        "htl garten", "altstädt", "mautsch", "hortus galabau",
        # PLZ 8 (Bayern Süd)
        "deutschmann galabau", "kaiser garten", "wittmann garten",
        "oberbauer", "kastrati", "schwaiger garten", "lechner garten",
        "bauer garten", "huber garten", "gartenbau müller",
        # PLZ 7 (Baden-Württemberg)
        "braun galabau", "albrecht bühler", "lutz + riepert",
        "drautz galabau", "garcke", "hartmann garten", "hill garten",
        "bodensee galabau", "schwaben galabau",
        # PLZ 6 (Hessen/Saarland/Rhein-Neckar)
        "winkler garten", "gramenz", "burkhardt garten", "seitz garten",
        "bach galabau", "garten moser", "scherer garten",
        # PLZ 5 (Köln/Bonn/Aachen)
        "nagelschmitz", "klara gmbh", "liesenberg", "galabau ertuerk",
        "galabau sieger", "bauer courth", "pankraz", "zimmermann garten",
        "galabau salber", "galabau weirauch", "loboda",
        # PLZ 4 (NRW Ruhrgebiet)
        "gehrken", "plum garten", "scheidtmann", "peter rose",
        "becher galabau", "lemp garten", "möhle", "groppe",
        "kaika garten", "eumann", "wieschen", "blanik",
        "terfruechte", "klein garten", "riesop",
        # PLZ 3 (Hannover/Braunschweig/Kassel)
        "janisch galabau", "keller tersch", "drewes garten",
        "rebohl", "steuber", "benning garten", "broxtermann",
        "friedrichs garten", "vornkahl", "kleyböcker",
        # PLZ 2 (Hamburg/SH/Niedersachsen Nord)
        "labarre", "heino harms", "schnoor garten", "erwin rumpf",
        "darger garten", "backhaus garten", "grewe galabau",
        "meykopff", "rieckhof", "tietjen garten",
        # PLZ 1 (Berlin/Brandenburg/MV)
        "häntsch", "scharf garten", "fehmer", "dalhoff",
        "gebrüder pfeil", "rohrbeck garten", "arkadia garten",
        "binner garten", "röder garten", "lubitz garten",
        # PLZ 0 (Sachsen/Sachsen-Anhalt)
        "baum & garten", "exact galabau", "bleyer garten",
        "hofmann garten", "schönherr garten", "henkler",
        "riede garten", "klemm garten", "winkler garten",
    ]

    # --- C-LEADS: Landschaftsarchitekten (indirekt, kaufen nicht direkt) ---
    if "landschaftsarchitekt" in kategorie:
        return "C"

    # --- A-LEADS: Kommune mit Phone+Email, oder GaLaBau-Betrieb mit Wegebau-Fokus ---
    if "kommune" in kategorie:
        if has_phone and (has_email or has_contact):
            return "A"
        if has_phone or has_email:
            return "B"
        return "B"

    # Private GaLaBau-Betriebe
    if "privat" in kategorie:
        is_specialist = has_galabau_relevance or any(kw in name for kw in high_relevance_names)

        if "großunternehmen" in kategorie:
            return "B"

        if is_specialist and has_phone and has_email:
            return "A"
        if is_specialist and (has_phone or has_email):
            return "B"
        if has_phone and has_email:
            return "B"
        if has_phone or has_email:
            return "B"
        return "C"

    # --- B-LEADS: Staatliche Gartenverwaltungen, Landesbetriebe ---
    if "behörde" in kategorie or "land" in kategorie:
        return "B"

    return "B"


def get_gespraechsaufhaenger(rec):
    """Generate a GaLaBau-specific conversation starter based on category."""
    kategorie = rec["kategorie"].lower()

    if "kommune" in kategorie:
        return "Frostschäden an Parkwegen und öffentlichen Flächen schnell beheben — ARM Kaltasphalt als Sofortlösung"
    elif "privat" in kategorie:
        return "Schnelle Asphalt-Reparatur für Einfahrten, Wege und Terrassen — ARM Kaltasphalt, sofort verarbeitbar"
    elif "behörde" in kategorie or "land" in kategorie:
        return "Kaltasphalt für Sofort-Reparaturen an Wegen und Plätzen in öffentlichen Grünanlagen"
    elif "landschaftsarchitekt" in kategorie:
        return "ARM Kaltasphalt als Spezifikationsoption für schnelle Wegesanierung"
    else:
        return "Schnelle Frostschaden-Reparatur an Wegen und Flächen mit ARM Kaltasphalt"


# ---------------------------------------------------------------------------
# 4. NAME SPLITTING
# ---------------------------------------------------------------------------

def split_contact_name(kontaktperson):
    """Split contact person into first name and last name."""
    if not kontaktperson or kontaktperson in ["–", "-", ""]:
        return "", ""

    name = re.sub(r'\(.*?\)', '', kontaktperson).strip()
    for title in ["Dr. Ing.", "Dr.-Ing.", "Dr.", "Dipl.-Ing.", "Prof.",
                   "Herr", "Frau", "Ing."]:
        name = name.replace(title, "").strip()

    parts = name.split()
    if len(parts) == 0:
        return "", ""
    elif len(parts) == 1:
        return "", parts[0]
    else:
        return parts[0], " ".join(parts[1:])


def extract_role(kontaktperson):
    """Extract role/position from contact person string."""
    if not kontaktperson:
        return ""
    match = re.search(r'\(([^)]+)\)', kontaktperson)
    if match:
        role = match.group(1)
        role_map = {
            "GF": "Geschäftsführer",
            "Vertretung": "Vertretungsberechtigter",
            "vertretungsberechtigt": "Vertretungsberechtigter",
            "Amtsleitung": "Amtsleiter",
            "Amtsleiter": "Amtsleiter",
            "Fachbereichsleiter": "Fachbereichsleiter",
            "Direktor": "Direktor",
            "Kontakt": "Kontakt",
            "Inhaber": "Inhaber",
        }
        return role_map.get(role, role)
    return ""


def determine_abteilung(rec):
    """Determine department based on GaLaBau category."""
    kategorie = rec["kategorie"].lower()
    name = rec["name"].lower()

    if "grünfläche" in name or "gartenamt" in name or "grünamt" in name:
        return "Grünflächenamt"
    elif "garten" in name and ("amt" in name or "verwaltung" in name):
        return "Gartenverwaltung"
    elif "landschaftsarchitekt" in kategorie or "planungsbüro" in name:
        return "Planung"
    elif "kommune" in kategorie:
        return "Grünflächenpflege"
    elif "behörde" in kategorie:
        return "Gartenverwaltung"
    elif "galabau" in name or "garten" in name or "landschaft" in name:
        return "GaLaBau"
    else:
        return "GaLaBau"


# ---------------------------------------------------------------------------
# 5. EXPORT
# ---------------------------------------------------------------------------

def export_crm_csv(records, filepath):
    """Export leads in Salesforce CRM import format."""
    fieldnames = [
        "Name", "FirstName", "LastName", "PostalCode", "City",
        "Email", "Phone", "MobilePhone", "Company", "Title",
        "IBS_SC_Position__c", "IBS_SC_Abteilungen__c"
    ]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()

        for rec in records:
            first, last = split_contact_name(rec["kontaktperson"])
            if not last:
                last = rec["name"][:40]

            writer.writerow({
                "Name": f"{first} {last}".strip() if first else last,
                "FirstName": first,
                "LastName": last,
                "PostalCode": rec["plz"],
                "City": rec["ort"],
                "Email": rec["email"],
                "Phone": rec["telefon"],
                "MobilePhone": "",
                "Company": rec["name"],
                "Title": f"{rec['priority']}-Lead ARM GaLaBau",
                "IBS_SC_Position__c": extract_role(rec["kontaktperson"]),
                "IBS_SC_Abteilungen__c": determine_abteilung(rec),
            })

    print(f"CRM-Import CSV exportiert: {filepath}")


def export_anrufliste_csv(records, filepath):
    """Export prioritized call list."""
    fieldnames = [
        "Priorität", "Firma", "Kategorie", "Ansprechpartner",
        "Telefon", "Email", "PLZ", "Ort", "Gesprächsaufhänger", "Notiz"
    ]

    priority_order = {"A": 0, "B": 1, "C": 2}
    sorted_records = sorted(records, key=lambda r: (
        priority_order.get(r["priority"], 9),
        r["ort"]
    ))

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()

        for rec in sorted_records:
            writer.writerow({
                "Priorität": rec["priority"],
                "Firma": rec["name"],
                "Kategorie": rec["kategorie"],
                "Ansprechpartner": rec["kontaktperson"],
                "Telefon": rec["telefon"],
                "Email": rec["email"],
                "PLZ": rec["plz"],
                "Ort": rec["ort"],
                "Gesprächsaufhänger": get_gespraechsaufhaenger(rec),
                "Notiz": rec.get("leistungen", ""),
            })

    print(f"Anrufliste CSV exportiert: {filepath}")


def export_excel(records, filepath):
    """Export comprehensive Excel with A/B/C tabs."""
    if not HAS_OPENPYXL:
        print("Excel-Export übersprungen (openpyxl nicht installiert)")
        return

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fills = {
        "A": PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),  # dark green
        "B": PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),  # dark orange
        "C": PatternFill(start_color="37474F", end_color="37474F", fill_type="solid"),  # dark grey
        "Übersicht": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),  # green (GaLaBau)
    }
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    columns = [
        ("Firma", 40), ("Kategorie", 30), ("Ansprechpartner", 25),
        ("Position", 20), ("Telefon", 22), ("Email", 35),
        ("Straße", 30), ("PLZ", 8), ("Ort", 18), ("Bundesland", 15),
        ("Leistungen", 50), ("Gesprächsaufhänger", 55),
    ]

    priority_order = {"A": 0, "B": 1, "C": 2}
    sorted_records = sorted(records, key=lambda r: (
        priority_order.get(r["priority"], 9),
        r["ort"]
    ))

    # Create overview sheet
    ws_overview = wb.active
    ws_overview.title = "Übersicht"
    ws_overview.append(["ARM GaLaBau Kampagne - Leadliste", "", "", f"Stand: Februar 2026"])
    ws_overview.append([])
    ws_overview.append(["Priorität", "Anzahl", "Beschreibung"])

    a_count = sum(1 for r in records if r["priority"] == "A")
    b_count = sum(1 for r in records if r["priority"] == "B")
    c_count = sum(1 for r in records if r["priority"] == "C")

    ws_overview.append(["A-Leads", a_count,
                         "Kommunale Grünflächenämter + kleine/mittlere GaLaBau-Betriebe mit Wegebau-Fokus"])
    ws_overview.append(["B-Leads", b_count,
                         "Größere GaLaBau-Ketten + Staatliche Gartenverwaltungen"])
    ws_overview.append(["C-Leads", c_count,
                         "Landschaftsarchitekten + Einträge ohne Kontaktdaten"])
    ws_overview.append(["GESAMT", a_count + b_count + c_count, ""])
    ws_overview.append([])
    ws_overview.append(["Angebot: Kaufe 2 Paletten (2x48 Sack), davon 24 Sack ohne Berechnung!"])
    ws_overview.append(["Laufzeit: Ende März (Option Ende April)"])

    # Style overview
    for cell in ws_overview[1]:
        if cell.value:
            cell.font = Font(bold=True, size=14)
    for cell in ws_overview[3]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = header_fills["Übersicht"]
            cell.font = header_font
    ws_overview.column_dimensions['A'].width = 15
    ws_overview.column_dimensions['B'].width = 10
    ws_overview.column_dimensions['C'].width = 70

    # Create A/B/C tabs
    for priority in ["A", "B", "C"]:
        priority_records = [r for r in sorted_records if r["priority"] == priority]
        label = {"A": "A-Leads (Sofort)", "B": "B-Leads (Zweite Welle)", "C": "C-Leads (Optional)"}
        ws = wb.create_sheet(title=label[priority])

        # Header row
        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fills[priority]
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width

        # Data rows
        for row_idx, rec in enumerate(priority_records, 2):
            values = [
                rec["name"],
                rec["kategorie"],
                rec["kontaktperson"],
                extract_role(rec["kontaktperson"]),
                rec["telefon"],
                rec["email"],
                rec["strasse"],
                rec["plz"],
                rec["ort"],
                rec["bundesland"],
                rec["leistungen"],
                get_gespraechsaufhaenger(rec),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Freeze header row
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}1"

    wb.save(filepath)
    print(f"Excel exportiert: {filepath}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("ARM GaLaBau Kampagne - Leadliste Generator")
    print("=" * 60)

    # Step 1: Extract data from all configured PLZ regions
    print("\n[1/5] Daten extrahieren...")
    all_records = []

    for region_key, region_cfg in PLZ_REGIONS.items():
        print(f"\n  --- {region_cfg['label']} ---")

        # Claude report
        if region_cfg["claude"]:
            claude_file = SCRIPT_DIR / region_cfg["claude"]
            if claude_file.exists():
                claude_records = parse_claude_markdown_tables(claude_file)
                print(f"  Claude-Report:  {len(claude_records)} Einträge")
                all_records.extend(claude_records)
            else:
                print(f"  Claude-Report:  Datei nicht gefunden ({claude_file.name})")

    print(f"\n  Gesamt (roh):   {len(all_records)} Einträge aus {len(PLZ_REGIONS)} Regionen")

    # Step 1b: Clean up data
    for rec in all_records:
        if rec["plz"] and "x" in rec["plz"].lower():
            rec["plz"] = ""
        rec["ort"] = re.sub(r'\*+', '', rec["ort"]).strip()
        rec["ort"] = re.sub(r'^\d+[x]*\s+', '', rec["ort"]).strip()
        if rec["email"] and "@" not in rec["email"]:
            rec["email"] = ""
        if rec["kontaktperson"]:
            non_person = ["(keine person", "(kein ap)", "sekretariat"]
            if any(np in rec["kontaktperson"].lower() for np in non_person):
                rec["kontaktperson"] = ""
        if not rec["bundesland"] and rec["plz"]:
            rec["bundesland"] = detect_bundesland(rec["plz"])

    # Step 2: Deduplicate
    print("\n[2/5] Deduplizieren...")
    unique_records = deduplicate(all_records)
    print(f"  Nach Dedup:     {len(unique_records)} Einträge")
    print(f"  Entfernt:       {len(all_records) - len(unique_records)} Duplikate")

    # Step 3: Prioritize
    print("\n[3/5] Priorisierung (A/B/C)...")
    for rec in unique_records:
        rec["priority"] = categorize_priority(rec)

    a_leads = [r for r in unique_records if r["priority"] == "A"]
    b_leads = [r for r in unique_records if r["priority"] == "B"]
    c_leads = [r for r in unique_records if r["priority"] == "C"]

    print(f"  A-Leads (Sofort anrufen):    {len(a_leads)}")
    print(f"  B-Leads (Zweite Welle):      {len(b_leads)}")
    print(f"  C-Leads (Optional):          {len(c_leads)}")

    # Breakdown by Bundesland
    bundeslaender = {}
    for rec in unique_records:
        bl = rec.get("bundesland", "Unbekannt") or "Unbekannt"
        bundeslaender[bl] = bundeslaender.get(bl, 0) + 1
    print("\n  Verteilung nach Bundesland:")
    for bl, count in sorted(bundeslaender.items()):
        print(f"    {bl:<25} {count}")

    # Step 4: Export CRM CSV
    print("\n[4/5] CRM-Import CSV exportieren...")
    crm_path = OUTPUT_DIR / "GaLaBau_CRM_Import_Leads.csv"
    export_crm_csv(unique_records, crm_path)

    # Step 5: Export Anrufliste
    print("\n[5/5] Anrufliste + Excel exportieren...")
    call_path = OUTPUT_DIR / "GaLaBau_Anrufliste_Priorisiert.csv"
    export_anrufliste_csv(unique_records, call_path)

    excel_path = OUTPUT_DIR / "GaLaBau_Leadliste_Komplett.xlsx"
    export_excel(unique_records, excel_path)

    # Summary
    print("\n" + "=" * 60)
    print("FERTIG!")
    print("=" * 60)
    print(f"\nDateien:")
    print(f"  1. {crm_path.name:<45} (Salesforce CRM Import)")
    print(f"  2. {call_path.name:<45} (Priorisierte Anrufliste)")
    if HAS_OPENPYXL:
        print(f"  3. {excel_path.name:<45} (Excel mit A/B/C Tabs)")

    print(f"\n--- Top 10 A-Leads ---")
    for i, rec in enumerate(sorted(a_leads, key=lambda r: r["ort"]), 1):
        contact = rec["kontaktperson"][:30] if rec["kontaktperson"] else "(kein AP)"
        print(f"  {i:2d}. {rec['name'][:45]:<45} | {rec['ort']:<15} | {contact}")
        if i >= 10:
            break


if __name__ == "__main__":
    main()
