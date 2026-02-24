#!/usr/bin/env python3
"""Export ARM Kampagne Leadliste als Excel mit Tier-Tabs"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def has_contact(ap):
    if not ap: return False
    ap = ap.strip()
    return ap and ap not in ('', '-', 'nicht gefunden')

# Load data
with open('ARM_Kampagne_Gesamtliste.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f, delimiter=';')
    fieldnames = reader.fieldnames
    rows = list(reader)

prio_field = fieldnames[0]

# Split into tiers
tier1, tier2_kommun, tier2_other, b_leads, c_leads = [], [], [], [], []

for r in rows:
    prio = r.get(prio_field, '')
    ap = r.get('Ansprechpartner', '')
    cat = r.get('Kategorie', '')

    if prio == 'A':
        if has_contact(ap):
            tier1.append(r)
        elif cat == 'Kommune':
            tier2_kommun.append(r)
        else:
            tier2_other.append(r)
    elif prio == 'B':
        b_leads.append(r)
    else:
        c_leads.append(r)

tier1.sort(key=lambda r: (r.get('Kategorie',''), r.get('Ort','')))
tier2_kommun.sort(key=lambda r: r.get('Ort',''))
tier2_other.sort(key=lambda r: (r.get('Kategorie',''), r.get('Ort','')))
b_leads.sort(key=lambda r: (r.get('Kategorie',''), r.get('Ort','')))
c_leads.sort(key=lambda r: (r.get('Kategorie',''), r.get('Ort','')))

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill_green = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
header_fill_orange = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
header_fill_blue = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_fill_gray = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
kommun_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
galabau_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
strassenbau_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

columns = [
    ("Prio", prio_field, 6),
    ("Firma", "Firma", 45),
    ("Kategorie", "Kategorie", 22),
    ("Ansprechpartner", "Ansprechpartner", 25),
    ("Telefon", "Telefon", 22),
    ("Email", "Email", 35),
    ("PLZ", "PLZ", 8),
    ("Ort", "Ort", 25),
    ("Notiz", "Notiz", 35),
    ("Quelle", "Quelle", 10),
]

def write_sheet(ws, data, sheet_fill):
    for col_idx, (name, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = sheet_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, r in enumerate(data, 2):
        cat = r.get('Kategorie', '')
        for col_idx, (_, field, _) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(field, ''))
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=(col_idx in (2, 9)))
            if 'Kommune' in cat:
                cell.fill = kommun_fill
            elif 'GaLaBau' in cat:
                cell.fill = galabau_fill
            elif 'Stra' in cat:
                cell.fill = strassenbau_fill

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data)+1}"

wb = openpyxl.Workbook()

# Tab 1: Tier 1
ws1 = wb.active
ws1.title = f"Tier1 Anrufbereit ({len(tier1)})"
write_sheet(ws1, tier1, header_fill_green)

# Tab 2: Tier 2 Kommunen
ws2 = wb.create_sheet(f"Tier2 Kommunen o.AP ({len(tier2_kommun)})")
write_sheet(ws2, tier2_kommun, header_fill_orange)

# Tab 3: Tier 2 Privat
ws3 = wb.create_sheet(f"Tier2 Privat o.AP ({len(tier2_other)})")
write_sheet(ws3, tier2_other, header_fill_orange)

# Tab 4: B-Leads
ws4 = wb.create_sheet(f"B-Leads ({len(b_leads)})")
write_sheet(ws4, b_leads, header_fill_blue)

# Tab 5: C-Leads
ws5 = wb.create_sheet(f"C-Leads ({len(c_leads)})")
write_sheet(ws5, c_leads, header_fill_gray)

# Tab 6: Zusammenfassung
ws6 = wb.create_sheet("Zusammenfassung")
ws6.sheet_properties.tabColor = "2F5496"
for c in ['A','B','C','D','E']:
    ws6.column_dimensions[c].width = [35, 12, 12, 12, 12]['ABCDE'.index(c)]

summary = [
    ["ARM Kaltasphalt - Kampagnen-Leadliste"],
    ["Stand: 24.02.2026"],
    [""],
    ["SEGMENT", "A-LEADS", "MIT AP", "QUOTE", "OHNE AP"],
    ["Kommune", 395, 306, "77%", 89],
    ["Privat (GaLaBau)", 139, 119, "86%", 20],
    ["Privat (Strassenbau)", 87, 38, "44%", 49],
    ["GESAMT A-Leads", 621, 463, "75%", 158],
    [""],
    ["TIER-AUFTEILUNG", "ANZAHL", "", "", ""],
    [f"Tier 1 - Sofort anrufbar (AP+Tel)", len(tier1)],
    [f"Tier 2 - Kommunen ohne AP", len(tier2_kommun)],
    [f"Tier 2 - Privat ohne AP", len(tier2_other)],
    ["B-Leads (Nachfass)", len(b_leads)],
    ["C-Leads (Langfrist)", len(c_leads)],
    ["GESAMT", len(rows)],
    [""],
    ["LEGENDE"],
    ["Blau hinterlegt = Kommune"],
    ["Gruen hinterlegt = GaLaBau"],
    ["Orange hinterlegt = Strassenbau"],
    [""],
    ["HINWEISE"],
    ["- Tier 1 Leads sind sofort telefonisch kontaktierbar"],
    ["- Tier 2 Kommunen: Bauhof anrufen, nach Leiter fragen"],
    ["- Tier 2 Privat: meist Strassenbaufirmen mit Datenqualitaetsproblemen"],
    ["- B-Leads eignen sich fuer E-Mail-Kampagne"],
    ["- CSV fuer CRM-Import: ARM_Kampagne_Gesamtliste.csv"],
]

for row_idx, row_data in enumerate(summary, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif row_idx == 2:
            cell.font = Font(size=11, italic=True)
        elif row_idx in (4, 10):
            cell.font = header_font
            cell.fill = header_fill_blue
        elif row_idx in (8, 16):
            cell.font = Font(bold=True, size=11)
        elif row_idx == 18:
            cell.font = Font(bold=True)
        elif row_idx == 19:
            cell.fill = kommun_fill
        elif row_idx == 20:
            cell.fill = galabau_fill
        elif row_idx == 21:
            cell.fill = strassenbau_fill

output = "ARM_Kampagne_Leadliste.xlsx"
wb.save(output)
print(f"Excel gespeichert: {output}")
print()
print(f"Tabs:")
print(f"  1. Tier 1 Anrufbereit:     {len(tier1)} Leads")
print(f"  2. Tier 2 Kommunen o.AP:   {len(tier2_kommun)} Leads")
print(f"  3. Tier 2 Privat o.AP:     {len(tier2_other)} Leads")
print(f"  4. B-Leads:                {len(b_leads)} Leads")
print(f"  5. C-Leads:                {len(c_leads)} Leads")
print(f"  6. Zusammenfassung")
print(f"  GESAMT:                    {len(rows)} Leads")
