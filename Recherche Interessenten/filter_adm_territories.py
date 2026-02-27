#!/usr/bin/env python3
"""
Filtert ARM_Kampagne_Gesamtliste.csv auf die Verkaufsgebiete der 4 ADM-Fachberater
und erzeugt:
  1. ARM_ADM_Gesamtliste.csv      — Alle Leads in ADM-Gebieten (mit Fachberater-Spalte)
  2. ARM_ADM_Kampagne.xlsx         — Excel mit Übersicht + Fachberater-Tabs
  3. ARM_ADM_CRM_Import.csv        — Salesforce-Import (nur kampagnenbereite Leads)

Verwendung:
  cd "Recherche Interessenten"
  python filter_adm_territories.py
"""

import csv
import io
import re
import os
import sys

# Windows UTF-8 Konsolen-Fix
if sys.stdout and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr and hasattr(sys.stderr, "buffer"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("FEHLER: openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "ARM_Kampagne_Gesamtliste.csv")
OUTPUT_CSV = os.path.join(SCRIPT_DIR, "ARM_ADM_Gesamtliste.csv")
OUTPUT_XLSX = os.path.join(SCRIPT_DIR, "ARM_ADM_Kampagne.xlsx")
OUTPUT_CRM = os.path.join(SCRIPT_DIR, "ARM_ADM_CRM_Import.csv")

# ---------------------------------------------------------------------------
# PLZ → Fachberater Zuordnung
# ---------------------------------------------------------------------------

TERRITORY_MAP = {
    # Jens Sackmann: 20-29
    20: "Jens Sackmann", 21: "Jens Sackmann", 22: "Jens Sackmann", 23: "Jens Sackmann",
    24: "Jens Sackmann", 25: "Jens Sackmann", 26: "Jens Sackmann", 27: "Jens Sackmann",
    28: "Jens Sackmann", 29: "Jens Sackmann",
    # André Grahn: 40-49, 50-53, 57-59
    40: "André Grahn", 41: "André Grahn", 42: "André Grahn", 43: "André Grahn",
    44: "André Grahn", 45: "André Grahn", 46: "André Grahn", 47: "André Grahn",
    48: "André Grahn", 49: "André Grahn",
    50: "André Grahn", 51: "André Grahn", 52: "André Grahn", 53: "André Grahn",
    57: "André Grahn", 58: "André Grahn", 59: "André Grahn",
    # Jens Lang: 70-79, 86-89
    70: "Jens Lang", 71: "Jens Lang", 72: "Jens Lang", 73: "Jens Lang",
    74: "Jens Lang", 75: "Jens Lang", 76: "Jens Lang", 77: "Jens Lang",
    78: "Jens Lang", 79: "Jens Lang",
    86: "Jens Lang", 87: "Jens Lang", 88: "Jens Lang", 89: "Jens Lang",
    # Daniel May: 80-85, 94
    80: "Daniel May", 81: "Daniel May", 82: "Daniel May", 83: "Daniel May",
    84: "Daniel May", 85: "Daniel May",
    94: "Daniel May",
    # Francesco Palese: 90-93, 95-97
    90: "Francesco Palese", 91: "Francesco Palese", 92: "Francesco Palese",
    93: "Francesco Palese",
    95: "Francesco Palese", 96: "Francesco Palese", 97: "Francesco Palese",
}

# Sortierreihenfolge für Fachberater
FACHBERATER_ORDER = ["Jens Sackmann", "André Grahn", "Jens Lang", "Daniel May", "Francesco Palese"]

# Priorität → Sortierrang
PRIO_ORDER = {"A": 0, "B": 1, "C": 2}

# Priorität → Salesforce Rating
PRIORITY_TO_RATING = {"A": "Hot", "B": "Warm", "C": "Cold"}

# Kategorie → Salesforce Industry (aus convert_arm_to_crm.py)
KATEGORIE_TO_INDUSTRY = {
    "Kommune": "Government",
    "Privat (GaLaBau)": "Construction",
    "Privat (Straßenbau)": "Construction",
}

# Salesforce CRM Header (aus convert_arm_to_crm.py)
CRM_HEADERS = [
    "First Name", "Last Name", "Company", "Title", "Phone", "Email",
    "Lead Status", "Rating", "Street", "City", "State/Province",
    "Zip/Postal Code", "Country", "Website", "No. Of Employees",
    "Annual Revenue", "Lead Source", "Industry", "Description"
]

# Salutations to strip (aus convert_arm_to_crm.py)
SALUTATIONS = {"herr", "frau", "dr.", "prof.", "ing.", "dipl.-ing."}


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def get_fachberater(plz):
    """PLZ (str) → Fachberater-Name oder None."""
    plz = plz.strip()
    if len(plz) < 2 or not plz[:2].isdigit():
        return None
    return TERRITORY_MAP.get(int(plz[:2]))


def has_contact(ap):
    """Prüft ob ein verwertbarer Ansprechpartner vorhanden ist (aus export_excel.py)."""
    if not ap:
        return False
    ap = ap.strip()
    return ap and ap not in ("", "-", "nicht gefunden")


def classify_readiness(row, prio_field):
    """Klassifiziert Lead-Bereitschaft: 'bereit' / 'anrufbar' / 'recherche'."""
    ap = row.get("Ansprechpartner", "").strip()
    tel = row.get("Telefon", "").strip()
    email = row.get("Email", "").strip()

    if has_contact(ap) and tel and email:
        return "bereit"       # AP + Tel + Email → voll kampagnenbereit
    elif has_contact(ap) and tel:
        return "anrufbar"     # AP + Tel (ohne Email) → anrufbar
    else:
        return "recherche"    # Fehlende Kontaktdaten


def parse_ansprechpartner(name_raw):
    """Parse Ansprechpartner into (first_name, last_name, title). Aus convert_arm_to_crm.py."""
    if not name_raw or not name_raw.strip():
        return "", "", ""

    name = name_raw.strip()

    title = ""
    paren_match = re.search(r'\(([^)]+)\)', name)
    if paren_match:
        paren_content = paren_match.group(1).strip()
        non_titles = {"Kontakt", "Vertretung", "Sample"}
        if paren_content not in non_titles:
            title = paren_content
        name = name[:paren_match.start()].strip()

    parts = name.split()
    if not parts:
        return "", "", title

    cleaned_parts = []
    for part in parts:
        if part.lower().rstrip(".") + "." in SALUTATIONS or part.lower() in SALUTATIONS:
            continue
        cleaned_parts.append(part)

    if not cleaned_parts:
        cleaned_parts = parts

    if len(cleaned_parts) == 1:
        return "", cleaned_parts[0], title
    else:
        return cleaned_parts[0], " ".join(cleaned_parts[1:]), title


# ---------------------------------------------------------------------------
# Laden und Filtern
# ---------------------------------------------------------------------------

def load_and_filter():
    """Liest Gesamtliste, filtert auf ADM-Gebiete, fügt Fachberater hinzu."""
    with open(INPUT_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        fieldnames = reader.fieldnames
        all_rows = list(reader)

    # Detect BOM in Priorität field
    prio_field = fieldnames[0]  # Might be '\ufeffPriorität' or 'Priorität'

    filtered = []
    for row in all_rows:
        plz = row.get("PLZ", "").strip()
        fb = get_fachberater(plz)
        if fb:
            row["Fachberater"] = fb
            row["_prio"] = row.get(prio_field, "").strip()
            row["_readiness"] = classify_readiness(row, prio_field)
            filtered.append(row)

    # Sortierung: Fachberater → Priorität (A→B→C) → Ort
    filtered.sort(key=lambda r: (
        FACHBERATER_ORDER.index(r["Fachberater"]) if r["Fachberater"] in FACHBERATER_ORDER else 99,
        PRIO_ORDER.get(r["_prio"], 9),
        r.get("Ort", ""),
    ))

    return filtered, prio_field


# ---------------------------------------------------------------------------
# Export: Gefilterte CSV
# ---------------------------------------------------------------------------

def export_filtered_csv(rows, prio_field):
    """Erzeugt ARM_ADM_Gesamtliste.csv — nur kampagnenbereite Leads mit Fachberater-Spalte."""
    # Nur kampagnenbereite Leads (A bereit + B bereit)
    ready_rows = [r for r in rows
                  if (r["_prio"] == "A" and r["_readiness"] == "bereit")
                  or (r["_prio"] == "B" and r["_readiness"] == "bereit")]

    headers = [prio_field, "Firma", "Kategorie", "Ansprechpartner", "Telefon",
               "Email", "PLZ", "Ort", "Fachberater", "Gesprächsaufhänger", "Notiz", "Quelle"]

    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter=";", extrasaction="ignore")
        writer.writeheader()
        for row in ready_rows:
            out = dict(row)
            out[prio_field] = row["_prio"]
            writer.writerow(out)

    return len(ready_rows)


# ---------------------------------------------------------------------------
# Export: Excel mit Fachberater-Tabs
# ---------------------------------------------------------------------------

def export_excel(rows, prio_field):
    """Erzeugt ARM_ADM_Kampagne.xlsx — nur kampagnenbereite Leads (A bereit + B bereit)."""
    # Nur kampagnenbereite Leads (identisch mit CRM-Import)
    ready_rows = [r for r in rows
                  if (r["_prio"] == "A" and r["_readiness"] == "bereit")
                  or (r["_prio"] == "B" and r["_readiness"] == "bereit")]

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_blue = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill_green = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

    fill_a = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Grün = A (Hot)
    fill_b = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")   # Hellblau = B (Warm)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    columns = [
        ("Prio", "_prio", 6),
        ("Firma", "Firma", 45),
        ("Kategorie", "Kategorie", 22),
        ("Ansprechpartner", "Ansprechpartner", 25),
        ("Telefon", "Telefon", 22),
        ("Email", "Email", 35),
        ("PLZ", "PLZ", 8),
        ("Ort", "Ort", 25),
        ("Gesprächsaufhänger", "Gesprächsaufhänger", 40),
        ("Notiz", "Notiz", 30),
        ("Quelle", "Quelle", 10),
        ("Status", None, 18),  # Leere Spalte für Anruf-Tracking
    ]

    def write_leads_sheet(ws, data, sheet_fill):
        """Schreibt Lead-Daten in ein Worksheet."""
        for col_idx, (name, _, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = sheet_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, r in enumerate(data, 2):
            prio = r.get("_prio", "")

            for col_idx, (_, field, _) in enumerate(columns, 1):
                value = r.get(field, "") if field else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (2, 9, 10)))

                if prio == "A":
                    cell.fill = fill_a
                elif prio == "B":
                    cell.fill = fill_b

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data) + 1}"

    # --- Übersicht-Tab ---
    ws_overview = wb.active
    ws_overview.title = "Übersicht"
    ws_overview.sheet_properties.tabColor = "2F5496"

    for col, width in [("A", 25), ("B", 12), ("C", 12), ("D", 12)]:
        ws_overview.column_dimensions[col].width = width

    # Stats pro Fachberater (nur kampagnenbereite)
    stats = {}
    for fb in FACHBERATER_ORDER:
        fb_rows = [r for r in ready_rows if r["Fachberater"] == fb]
        a_rows = [r for r in fb_rows if r["_prio"] == "A"]
        b_rows = [r for r in fb_rows if r["_prio"] == "B"]
        stats[fb] = {
            "gesamt": len(fb_rows),
            "a": len(a_rows),
            "b": len(b_rows),
        }

    totals = {k: sum(stats[fb][k] for fb in FACHBERATER_ORDER)
              for k in ["gesamt", "a", "b"]}

    summary_data = [
        ["ARM Kampagne — Kampagnenbereite Leads"],
        ["Stand: 27.02.2026"],
        ["Nur Leads mit AP + Telefon + Email (sofort kontaktierbar)"],
        [""],
        ["Fachberater", "Gesamt", "A (Hot)", "B (Warm)"],
    ]

    for fb in FACHBERATER_ORDER:
        s = stats[fb]
        summary_data.append([fb, s["gesamt"], s["a"], s["b"]])

    summary_data.append(["TOTAL", totals["gesamt"], totals["a"], totals["b"]])
    summary_data.extend([
        [""],
        ["FARBKODIERUNG"],
        ["Grün = A-Lead (Hot)"],
        ["Hellblau = B-Lead (Warm)"],
        [""],
        ["HINWEISE"],
        ["Identischer Inhalt wie CRM-Import (Salesforce)"],
        ["Leere Status-Spalte für Anruf-Tracking"],
    ])

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_overview.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 2:
                cell.font = Font(size=11, italic=True)
            elif row_idx == 3:
                cell.font = Font(size=10, italic=True)
            elif row_idx == 5:
                cell.font = header_font
                cell.fill = header_fill_blue
            elif row_idx == 10:  # TOTAL
                cell.font = Font(bold=True, size=11)
            elif row_idx == 12:  # FARBKODIERUNG Header
                cell.font = Font(bold=True)
            elif row_idx == 13:
                cell.fill = fill_a
            elif row_idx == 14:
                cell.fill = fill_b

    # --- Fachberater-Tabs ---
    for fb in FACHBERATER_ORDER:
        fb_rows = [r for r in ready_rows if r["Fachberater"] == fb]
        s = stats[fb]
        tab_name = f"{fb} ({s['gesamt']})"
        if len(tab_name) > 31:
            tab_name = tab_name[:31]
        ws = wb.create_sheet(tab_name)
        write_leads_sheet(ws, fb_rows, header_fill_green)

    wb.save(OUTPUT_XLSX)
    return len(ready_rows)


# ---------------------------------------------------------------------------
# Export: CRM-Import CSV (Salesforce-Format)
# ---------------------------------------------------------------------------

def export_crm_csv(rows):
    """Erzeugt ARM_ADM_CRM_Import.csv — nur kampagnenbereite Leads im Salesforce-Format."""
    # A bereit + B bereit
    crm_rows = [r for r in rows
                 if (r["_prio"] == "A" and r["_readiness"] == "bereit")
                 or (r["_prio"] == "B" and r["_readiness"] == "bereit")]

    with open(OUTPUT_CRM, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CRM_HEADERS)
        writer.writeheader()

        for row in crm_rows:
            first_name, last_name, title = parse_ansprechpartner(
                row.get("Ansprechpartner", "")
            )

            # Description mit Fachberater für Routing
            desc_parts = []
            desc_parts.append(f"Fachberater: {row.get('Fachberater', '')}")
            if row.get("Gesprächsaufhänger", "").strip():
                desc_parts.append(f"Gesprächsaufhänger: {row['Gesprächsaufhänger'].strip()}")
            if row.get("Notiz", "").strip():
                desc_parts.append(f"Notiz: {row['Notiz'].strip()}")
            if row.get("Quelle", "").strip():
                desc_parts.append(f"Quelle: {row['Quelle'].strip()}")
            description = " | ".join(desc_parts)

            kategorie = row.get("Kategorie", "").strip()

            crm_row = {
                "First Name": first_name,
                "Last Name": last_name or row.get("Firma", "").strip()[:40],
                "Company": row.get("Firma", "").strip(),
                "Title": title,
                "Phone": row.get("Telefon", "").strip(),
                "Email": row.get("Email", "").strip(),
                "Lead Status": "New",
                "Rating": PRIORITY_TO_RATING.get(row["_prio"], ""),
                "Street": "",
                "City": row.get("Ort", "").strip(),
                "State/Province": "",
                "Zip/Postal Code": row.get("PLZ", "").strip(),
                "Country": "Germany",
                "Website": "",
                "No. Of Employees": "",
                "Annual Revenue": "",
                "Lead Source": "ARM Kampagne",
                "Industry": KATEGORIE_TO_INDUSTRY.get(kategorie, kategorie),
                "Description": description,
            }

            writer.writerow(crm_row)

    return len(crm_rows)


# ---------------------------------------------------------------------------
# Report auf Konsole
# ---------------------------------------------------------------------------

def print_report(rows):
    """Gibt Zusammenfassung auf Konsole aus."""
    # Kampagnenbereite Leads
    ready = [r for r in rows
             if (r["_prio"] == "A" and r["_readiness"] == "bereit")
             or (r["_prio"] == "B" and r["_readiness"] == "bereit")]

    print("=" * 55)
    print("  ARM KAMPAGNE — KAMPAGNENBEREITE LEADS")
    print("  (Nur A+B mit AP + Telefon + Email)")
    print("=" * 55)
    print()

    header = f"{'Fachberater':<20} {'Gesamt':>7} {'A Hot':>7} {'B Warm':>7}"
    print(header)
    print("-" * len(header))

    grand = {"gesamt": 0, "a": 0, "b": 0}

    for fb in FACHBERATER_ORDER:
        fb_rows = [r for r in ready if r["Fachberater"] == fb]
        a_count = len([r for r in fb_rows if r["_prio"] == "A"])
        b_count = len([r for r in fb_rows if r["_prio"] == "B"])

        print(f"{fb:<20} {len(fb_rows):>7} {a_count:>7} {b_count:>7}")

        grand["gesamt"] += len(fb_rows)
        grand["a"] += a_count
        grand["b"] += b_count

    print("-" * len(header))
    print(f"{'TOTAL':<20} {grand['gesamt']:>7} {grand['a']:>7} {grand['b']:>7}")
    print()
    print(f"  Aus {len(rows)} Leads in ADM-Gebieten gefiltert")
    print(f"  {grand['gesamt']} kampagnenbereite Leads fuer CRM + Excel + CSV")
    print()


def print_files():
    """Gibt erzeugte Dateien aus."""
    print("  Erzeugte Dateien:")
    print(f"    1. {os.path.basename(OUTPUT_CSV)}")
    print(f"    2. {os.path.basename(OUTPUT_XLSX)}")
    print(f"    3. {os.path.basename(OUTPUT_CRM)}")
    print()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not os.path.exists(INPUT_FILE):
        print(f"FEHLER: Eingabedatei nicht gefunden: {INPUT_FILE}")
        sys.exit(1)

    rows, prio_field = load_and_filter()

    if not rows:
        print("FEHLER: Keine Leads in ADM-Gebieten gefunden.")
        sys.exit(1)

    # Report
    print_report(rows)

    # Exports
    csv_count = export_filtered_csv(rows, prio_field)
    xlsx_count = export_excel(rows, prio_field)
    crm_count = export_crm_csv(rows)

    print_files()
    print(f"  CSV:   {csv_count} Leads (Semikolon-Delimiter)")
    print(f"  Excel: {xlsx_count} Leads, {len(FACHBERATER_ORDER)} Fachberater-Tabs + Uebersicht")
    print(f"  CRM:   {crm_count} Leads (Salesforce-Format, Komma-Delimiter)")
    print()
    print("Alle 3 Dateien enthalten denselben Inhalt: nur kampagnenbereite Leads.")
    print("Fertig.")


if __name__ == "__main__":
    main()
