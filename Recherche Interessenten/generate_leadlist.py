#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ARM Kampagne - Leadliste Generator
Konsolidiert Daten aus beiden Research-Reports, priorisiert und exportiert
für CRM-Import und als Anrufliste.
"""

import csv
import io
import re
import os
from pathlib import Path

# Optional: openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("HINWEIS: openpyxl nicht installiert - Excel-Export wird übersprungen.")
    print("Installieren mit: pip install openpyxl\n")

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR

# PLZ region config: which source files to parse per region
PLZ_REGIONS = {
    "PLZ9": {
        "chatgpt": "deep-research-report chat GPT.md",
        "claude": "Recherche Tiefbauunternehmen PLZ 9 durch claude.md",
        "label": "PLZ 9 (Bayern/Thüringen)",
    },
    "PLZ0": {
        "chatgpt": None,  # No ChatGPT report for PLZ 0 yet
        "claude": "Recherche Tiefbauunternehmen PLZ 0 durch claude.md",
        "label": "PLZ 0 (Sachsen/Sachsen-Anhalt/Thüringen)",
    },
    "PLZ8": {
        "chatgpt": None,
        "claude": "Recherche Tiefbauunternehmen PLZ 8 durch claude.md",
        "label": "PLZ 8 (Bayern Süd - Oberbayern/Schwaben)",
    },
    "PLZ6": {
        "chatgpt": None,
        "claude": "Recherche Tiefbauunternehmen PLZ 6 durch claude.md",
        "label": "PLZ 6 (Hessen/Saarland/Rhein-Neckar)",
    },
    "PLZ7": {
        "chatgpt": None,
        "claude": "Recherche Tiefbauunternehmen PLZ 7 durch claude.md",
        "label": "PLZ 7 (Baden-Württemberg)",
    },
    "PLZ2": {
        "chatgpt": None,
        "claude": "Recherche Tiefbauunternehmen PLZ 2 durch claude.md",
        "label": "PLZ 2 (Hamburg/Schleswig-Holstein/Niedersachsen Nord/Bremen)",
    },
    "PLZ5": {
        "chatgpt": None,
        "claude": "Recherche Tiefbauunternehmen PLZ 5 durch claude.md",
        "label": "PLZ 5 (Köln/Bonn/Aachen/Koblenz/Trier)",
    },
    "PLZ4": {
        "chatgpt": None,
        "claude": "Recherche Tiefbauunternehmen PLZ 4 durch claude.md",
        "label": "PLZ 4 (NRW - Ruhrgebiet/Münsterland/Niederrhein)",
    },
    "PLZ1": {
        "chatgpt": None,
        "claude": "Recherche Tiefbauunternehmen PLZ 1 durch claude.md",
        "label": "PLZ 1 (Berlin/Brandenburg/Mecklenburg-Vorpommern)",
    },
    "PLZ3": {
        "chatgpt": None,
        "claude": "Recherche Tiefbauunternehmen PLZ 3 durch claude.md",
        "label": "PLZ 3 (Hannover/Braunschweig/Kassel/Göttingen/Magdeburg)",
    },
}


# ---------------------------------------------------------------------------
# 1. DATA EXTRACTION
# ---------------------------------------------------------------------------

def parse_chatgpt_csv(filepath):
    """Parse the CSV block from the ChatGPT deep-research report."""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    # Extract CSV block between ```csv and ```
    match = re.search(r"```csv\n(.*?)```", content, re.DOTALL)
    if not match:
        print("WARNUNG: Kein CSV-Block im ChatGPT-Report gefunden!")
        return []

    csv_text = match.group(1).strip()
    reader = csv.DictReader(io.StringIO(csv_text))
    records = []
    for row in reader:
        records.append({
            "source": "ChatGPT",
            "kategorie": row.get("Kategorie", "").strip(),
            "name": row.get("Name", "").strip(),
            "strasse": row.get("Straße", row.get("Stra\u00dfe", "")).strip(),
            "plz": row.get("PLZ", "").strip(),
            "ort": row.get("Ort", "").strip(),
            "bundesland": row.get("Bundesland", "").strip(),
            "zustaendigkeit": row.get("Zuständigkeit/Gebiet", row.get("Zust\u00e4ndigkeit/Gebiet", "")).strip(),
            "leistungen": row.get("Leistungen (Frostschaden-Bezug)", "").strip(),
            "kontaktperson": row.get("Kontaktperson (Rolle)", "").strip(),
            "telefon": row.get("Telefon", "").strip(),
            "email": row.get("E-Mail", "").strip(),
        })
    return records


def parse_claude_markdown_tables(filepath):
    """Parse markdown tables from the Claude research report."""
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    records = []
    current_section = ""
    in_table = False
    headers = []

    for i, line in enumerate(lines):
        stripped = line.strip()

        # Detect section headers
        if stripped.startswith("## 1."):
            current_section = "Kommune"
        elif stripped.startswith("## 2."):
            current_section = "Behörde"
        elif stripped.startswith("### Überregionale"):
            current_section = "Privat (Großunternehmen)"
        elif stripped.startswith("### Regionale"):
            current_section = "Privat (Mittelstand)"
        elif stripped.startswith("## 4."):
            current_section = "Ingenieurbüro"
        elif stripped.startswith("## Regionale Verteilung"):
            current_section = ""  # end of data

        # Detect table rows
        if stripped.startswith("|") and "---" not in stripped:
            cells = [c.strip().strip("*") for c in stripped.split("|")[1:-1]]

            if not in_table:
                # This is the header row
                headers = cells
                in_table = True
                continue

            if in_table and len(cells) >= 3:
                record = _parse_claude_table_row(cells, headers, current_section)
                if record:
                    records.append(record)
        elif in_table and not stripped.startswith("|"):
            in_table = False
            headers = []

    return records


def detect_bundesland(plz):
    """Detect Bundesland from PLZ prefix."""
    if not plz:
        return ""
    p = plz.strip()
    # Thüringen
    if p.startswith("98") or p.startswith("99"):
        return "Thüringen"
    if p.startswith("07") or p.startswith("046"):  # 04600 Altenburg
        return "Thüringen"
    # Sachsen
    if p.startswith("01") or p.startswith("02") or p.startswith("04") or \
       p.startswith("08") or p.startswith("09"):
        return "Sachsen"
    # Sachsen-Anhalt
    if p.startswith("06"):
        return "Sachsen-Anhalt"
    # Bayern (PLZ 8 + PLZ 9)
    if p.startswith("8") or p.startswith("9"):
        return "Bayern"
    # Baden-Württemberg (PLZ 7)
    if p.startswith("7"):
        return "Baden-Württemberg"
    # Hessen (PLZ 6 hauptsächlich + Nordhessen 34, 35, 36)
    if p.startswith("60") or p.startswith("61") or p.startswith("63") or \
       p.startswith("64") or p.startswith("65") or p.startswith("35") or \
       p.startswith("34") or p.startswith("36"):
        return "Hessen"
    # Saarland (PLZ 66)
    if p.startswith("66"):
        return "Saarland"
    # Rheinland-Pfalz (PLZ 67, parts of 55, 56, 54, 53)
    if p.startswith("67") or p.startswith("55") or p.startswith("56") or \
       p.startswith("54") or p.startswith("53"):
        return "Rheinland-Pfalz"
    # Baden-Württemberg (PLZ 68, 69, 7)
    if p.startswith("68") or p.startswith("69") or p.startswith("7"):
        return "Baden-Württemberg"
    # NRW (PLZ 4 + PLZ 5)
    if p.startswith("4") or p.startswith("5"):
        return "Nordrhein-Westfalen"
    # Mecklenburg-Vorpommern (PLZ 17, 18, 19)
    if p.startswith("17") or p.startswith("18") or p.startswith("19"):
        return "Mecklenburg-Vorpommern"
    # Berlin (PLZ 10-13)
    if p.startswith("10") or p.startswith("12") or p.startswith("13"):
        return "Berlin"
    # Brandenburg (PLZ 14-16)
    if p.startswith("14") or p.startswith("15") or p.startswith("16"):
        return "Brandenburg"
    # Hamburg (PLZ 20-22)
    if p.startswith("20") or p.startswith("21") or p.startswith("22"):
        return "Hamburg"
    # Schleswig-Holstein (PLZ 23-25)
    if p.startswith("23") or p.startswith("24") or p.startswith("25"):
        return "Schleswig-Holstein"
    # Bremen (PLZ 28)
    if p.startswith("28"):
        return "Bremen"
    # Sachsen-Anhalt (PLZ 39 - Magdeburg/Halberstadt/Stendal)
    if p.startswith("39"):
        return "Sachsen-Anhalt"
    # Niedersachsen (PLZ 26, 27, 29, 3)
    if p.startswith("26") or p.startswith("27") or p.startswith("29") or p.startswith("3"):
        return "Niedersachsen"
    return ""


def _parse_claude_table_row(cells, headers, section):
    """Convert a row from Claude's markdown tables into our unified format."""
    record = {
        "source": "Claude",
        "kategorie": section,
        "name": "",
        "strasse": "",
        "plz": "",
        "ort": "",
        "bundesland": "",
        "zustaendigkeit": "",
        "leistungen": "",
        "kontaktperson": "",
        "telefon": "",
        "email": "",
    }

    if section == "Kommune":
        # Headers: Ort | PLZ | Bezeichnung | Ansprechpartner | Telefon | E-Mail / Website
        if len(cells) >= 6:
            record["ort"] = cells[0]
            record["plz"] = cells[1]
            record["name"] = cells[2]
            record["kontaktperson"] = cells[3] if cells[3] != "–" else ""
            record["telefon"] = cells[4]
            email_web = cells[5]
            # Extract email from combined field
            emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', email_web)
            record["email"] = emails[0] if emails else ""
            record["leistungen"] = "Kommunaler Straßenunterhalt, Schlaglochbeseitigung"

    elif section == "Behörde":
        # Headers: Behörde | PLZ / Ort | Zuständigkeitsbereich | Telefon | E-Mail / Website
        if len(cells) >= 5:
            record["name"] = cells[0]
            plz_ort = cells[1]
            plz_match = re.match(r'(\d{5})\s+(.*)', plz_ort)
            if plz_match:
                record["plz"] = plz_match.group(1)
                record["ort"] = plz_match.group(2)
            record["zustaendigkeit"] = cells[2]
            record["telefon"] = cells[3]
            email_web = cells[4]
            emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', email_web)
            record["email"] = emails[0] if emails else ""
            record["leistungen"] = "Straßenbau/-unterhaltung Bundes-/Staatsstraßen"

    elif section.startswith("Privat"):
        if "Großunternehmen" in section:
            # Headers: Unternehmen | PLZ / Ort | Leistungsspektrum | GF / Vorstand | Kontakt
            if len(cells) >= 5:
                record["name"] = cells[0]
                plz_ort = cells[1]
                plz_match = re.match(r'(\d{5})\s+(.*)', plz_ort)
                if plz_match:
                    record["plz"] = plz_match.group(1)
                    record["ort"] = plz_match.group(2)
                else:
                    record["ort"] = plz_ort
                record["leistungen"] = cells[2]
                record["kontaktperson"] = cells[3] if cells[3] != "–" else ""
                kontakt = cells[4]
                phones = re.findall(r'[\d\s/\-·]+(?:\d{2,})', kontakt)
                if phones:
                    record["telefon"] = phones[0].strip().strip("·").strip()
                emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', kontakt)
                record["email"] = emails[0] if emails else ""
        else:
            # Mittelstand: Unternehmen | PLZ / Ort | Leistungsspektrum | Kontakt
            if len(cells) >= 4:
                record["name"] = cells[0]
                plz_ort = cells[1]
                plz_match = re.match(r'(\d{5})\s+(.*)', plz_ort)
                if plz_match:
                    record["plz"] = plz_match.group(1)
                    record["ort"] = plz_match.group(2)
                else:
                    record["ort"] = plz_ort
                record["leistungen"] = cells[2]
                kontakt = cells[3]
                phones = re.findall(r'[\d\s/\-]+(?:\d{2,})', kontakt)
                if phones:
                    record["telefon"] = phones[0].strip()
                emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', kontakt)
                record["email"] = emails[0] if emails else ""
        record["kategorie"] = "Privat (Straßenbau)"

    elif section == "Ingenieurbüro":
        # Headers: Büro | PLZ / Ort | Leistungsspektrum | Leitung / Inhaber | Kontakt
        if len(cells) >= 5:
            record["name"] = cells[0]
            plz_ort = cells[1]
            plz_match = re.match(r'(\d{5})\s+(.*)', plz_ort)
            if plz_match:
                record["plz"] = plz_match.group(1)
                record["ort"] = plz_match.group(2)
            else:
                record["ort"] = plz_ort
            record["leistungen"] = cells[2]
            record["kontaktperson"] = cells[3] if cells[3] != "–" else ""
            kontakt = cells[4]
            phones = re.findall(r'[\d\s/\-]+(?:\d{2,})', kontakt)
            if phones:
                record["telefon"] = phones[0].strip()
            emails = re.findall(r'[\w.+-]+@[\w.-]+\.\w+', kontakt)
            record["email"] = emails[0] if emails else ""

    # Skip empty records
    if not record["name"]:
        return None

    # Clean markdown artifacts from all fields
    for key in record:
        if isinstance(record[key], str):
            record[key] = record[key].replace("**", "").strip()

    # Auto-detect Bundesland from PLZ if not already set
    if not record["bundesland"] and record["plz"]:
        record["bundesland"] = detect_bundesland(record["plz"])

    return record


# ---------------------------------------------------------------------------
# 2. DEDUPLICATION
# ---------------------------------------------------------------------------

def normalize_name(name):
    """Normalize company name for dedup matching."""
    n = name.lower()
    # Remove common suffixes
    for suffix in ["gmbh & co. kg", "gmbh & co.kg", "gmbh", "se", "ag",
                    "e.k.", "gbr", "ohg", "kg"]:
        n = n.replace(suffix, "")
    # Remove special chars
    n = re.sub(r'[^a-zäöüß0-9\s]', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def extract_city_key(rec):
    """Extract a clean city name for city-based dedup of municipal orgs."""
    ort = rec.get("ort", "").strip()
    # Remove annotations like "(Lkr.)", "i.d.OPf." etc.
    ort = re.sub(r'\*+', '', ort)
    ort = re.sub(r'\s*\(.*?\)', '', ort)
    ort = ort.split("/")[0].strip()
    # Also try extracting from name: "Stadt Bamberg – ..." -> Bamberg
    name = rec.get("name", "")
    city_from_name = re.search(r'Stadt\s+(\w+)', name)
    if city_from_name:
        return city_from_name.group(1).lower().strip()
    return ort.lower().strip()


def _merge_records(winner, donor):
    """Merge contact data from donor into winner where winner has gaps."""
    for field in ["kontaktperson", "telefon", "email", "strasse",
                  "plz", "ort", "zustaendigkeit", "leistungen"]:
        if not winner[field] and donor[field]:
            winner[field] = donor[field]
    # Fix approximate PLZ: prefer real PLZ over "90xxx" style
    if winner["plz"] and "x" in winner["plz"].lower() and donor["plz"] and "x" not in donor["plz"].lower():
        winner["plz"] = donor["plz"]
    if donor["plz"] and "x" in donor["plz"].lower() and winner["plz"] and "x" not in winner["plz"].lower():
        pass  # winner already has good PLZ


def _contact_score(rec):
    """Score how much contact info a record has."""
    return sum([
        bool(rec["kontaktperson"] and rec["kontaktperson"] != "–"),
        bool(rec["telefon"]),
        bool(rec["email"]),
        bool(rec["strasse"]),
        bool(rec["plz"] and "x" not in rec["plz"].lower()),
    ])


# Manual mappings for known duplicates across the two sources.
# key = (city_lower) -> all records for that city+category get merged.
KOMMUNE_ALIASES = {
    # ChatGPT name fragments -> Claude name fragments (same org)
    "servicebetrieb öffentlicher raum": "sör",
    "bamberger service betrieb": "bamberg service",
    "tiefbauamt / eigenbetrieb ser": "tief- und landschaftsbau",
}

PRIVATE_ALIASES = {
    "firmengruppe max bögl": "max bögl bauservice",
    "josef rädlinger unternehmensgruppe": "josef rädlinger bauunternehmen",
    "rödl bau.gruppe": "rödl tiefbau",
}


def deduplicate(records):
    """Deduplicate records using multi-strategy matching."""
    # --- Strategy 1: City-based dedup for kommunale Tiefbauämter ---
    # Same city + same category (Kommune) = same organisation
    kommune_by_city = {}
    other_records = []

    for rec in records:
        kat = rec["kategorie"].lower()
        if "kommune" in kat:
            city_key = extract_city_key(rec)
            if not city_key:
                other_records.append(rec)
                continue
            if city_key in kommune_by_city:
                existing = kommune_by_city[city_key]
                if _contact_score(rec) > _contact_score(existing):
                    _merge_records(rec, existing)
                    kommune_by_city[city_key] = rec
                else:
                    _merge_records(existing, rec)
            else:
                kommune_by_city[city_key] = rec
        else:
            other_records.append(rec)

    # --- Strategy 2: Name-based dedup for non-kommune records ---
    seen = {}
    for rec in other_records:
        key = normalize_name(rec["name"])
        plz_prefix = rec["plz"][:3] if rec["plz"] and "x" not in rec["plz"][:3].lower() else ""
        dedup_key = f"{key}_{plz_prefix}"

        # Check aliases for known duplicates
        matched = False
        for alias_a, alias_b in PRIVATE_ALIASES.items():
            name_lower = rec["name"].lower()
            if alias_a in name_lower or alias_b in name_lower:
                alias_key = f"_alias_{alias_a}"
                if alias_key in seen:
                    existing = seen[alias_key]
                    if _contact_score(rec) > _contact_score(existing):
                        _merge_records(rec, existing)
                        seen[alias_key] = rec
                    else:
                        _merge_records(existing, rec)
                    matched = True
                else:
                    seen[alias_key] = rec
                    matched = True
                break

        if matched:
            continue

        # Also deduplicate entries where one has "xxx" PLZ and other has real PLZ
        # by trying to match on name alone if very similar
        name_only_key = key
        if dedup_key in seen:
            existing = seen[dedup_key]
            if _contact_score(rec) > _contact_score(existing):
                _merge_records(rec, existing)
                seen[dedup_key] = rec
            else:
                _merge_records(existing, rec)
        elif name_only_key in seen:
            existing = seen[name_only_key]
            if _contact_score(rec) > _contact_score(existing):
                _merge_records(rec, existing)
                seen[name_only_key] = rec
            else:
                _merge_records(existing, rec)
        else:
            seen[dedup_key] = rec
            seen[name_only_key] = rec  # Also index by name-only for cross-PLZ matching

    # Collect unique non-kommune records (avoid double-counting from name_only keys)
    unique_others = list({id(v): v for v in seen.values()}.values())

    result = list(kommune_by_city.values()) + unique_others
    return result


# ---------------------------------------------------------------------------
# 3. PRIORITIZATION
# ---------------------------------------------------------------------------

def categorize_priority(rec):
    """Assign A/B/C priority based on relevance for ARM campaign.

    A-Lead (~20-25): High relevance + good contact data -> call first
    B-Lead (~25-35): Medium relevance or missing contact data -> second wave
    C-Lead (rest):   Low direct relevance (engineering offices, Autobahn GmbH)
    """
    kategorie = rec["kategorie"].lower()
    name = rec["name"].lower()
    leistungen = rec["leistungen"].lower()
    has_contact = bool(rec["kontaktperson"] and rec["kontaktperson"] not in ["–", "-", ""])
    has_phone = bool(rec["telefon"])
    has_email = bool(rec["email"] and "@" in rec["email"])

    # Keywords indicating high frost-damage repair relevance
    frost_keywords = ["frostschad", "schlagloch", "instandsetzung", "sanierung",
                      "unterhalt", "reparatur", "deckeninstandsetzung",
                      "ausbesserung", "schadstelle", "spritzmaschine",
                      "aufgrabung"]
    has_frost_relevance = any(kw in leistungen for kw in frost_keywords)

    # Companies known for small-scale repair / municipal maintenance
    high_relevance_names = [
        # PLZ 9 (Bayern/Thüringen)
        "sar ", "sar straßen", "fuchs tiefbau", "sturm tiefbau",
        "hubert", "franken-asphalt", "donauasphalt", "nagler",
        "schill & geiger", "bayerische asphalt und umbauten",
        # PLZ 0 (Sachsen/Sachsen-Anhalt) - regional repair specialists
        "kirchner straßen", "baucom", "sbl straßenbau", "stra-ti",
        "vstr vogtland", "nickol bau", "strabau gmbh sangerhausen",
        "mitteldeutsche straßenbau", "chemnitzer verkehrsbau",
        "hastra-service", "ostthüringer straßenbau",
        "thiendorfer", "sbi straßen",
        # PLZ 8 (Bayern Süd - Oberbayern/Schwaben) - regional repair specialists
        "kutter", "arge asphalt", "swietelsky-faber",
        "held & francke", "tschuda bau", "georg maier tiefbau",
        "höfler straßen", "satzinger", "schmidbauer tiefbau",
        "dallmayr tief", "assner straßen", "gierl bau",
        "bodensee asphaltbau", "allgäuer straßenbau",
        "kemmer straßen", "heinrich lohr", "schmid straßen",
        # PLZ 6 (Hessen/Saarland/Rhein-Neckar) - regional repair specialists
        "peter gross", "dittgen bau", "giorgetti", "wilhelm faber",
        "colas rhein-main", "adam bau", "jakob bau", "jöst",
        "rohde & lie", "mohr straßen", "heinrich leser",
        "schenk & leuker", "helming", "schneider bau speyer",
        "streib", "backes bau",
        # PLZ 7 (Baden-Württemberg) - regional repair specialists
        "wolff & müller", "gottlob rommel", "g. kraft straßenbau",
        "epple tiefbau", "karl strohmaier", "friedrich geiger",
        "heinrich schlegel", "konrad bau", "reif bau",
        "ed. scherer", "storz verkehrswegebau", "schleith",
        "max wild", "kemmer straßen",
        # PLZ 5 (Köln/Bonn/Aachen/Koblenz/Trier) - regional repair specialists
        "kremer straßen", "linden gmbh", "siebengebirge asphalt",
        "schieß straßen", "heitkamp erd", "schmitz gmbh",
        "amos straßen", "kirchhoff straßen", "himmelmann straßen",
        "heckelmann", "schreinemacher", "eifel-bau",
        "modenbach", "heinrich betz", "august mainka",
        "karl gemünden",
        # PLZ 2 (Hamburg/SH/Niedersachsen Nord/Bremen) - regional repair specialists
        "stolz straßen", "bielfeldt & berg", "straßen- und tiefbau elmshorn",
        "johannsen & sohn", "straßen- und tiefbau lübeck", "arp straßen",
        "hillmer straßen", "tewes straßenbau", "hinrich söhnholz",
        "cordes straßen", "wesermarsch asphalt", "j. martens",
        "ehlers & pfuhl", "thielen straßen", "otto knüppel",
        "rohwedder + korth", "heinrich karstens", "reimer bau",
        # PLZ 4 (NRW Ruhrgebiet/Münsterland/Niederrhein) - regional repair specialists
        "heitkamp erd", "spieker straßen", "heckmann bau", "arge asphalt ruhr",
        "friedrich freund", "drees & hüsmann", "fritz spieker", "jopp gmbh",
        "middendorf gmbh", "kuhlmann gmbh", "niemeier gmbh", "stratmann gmbh",
        "tecklenborg", "balke-dürr", "wulfmeyer", "scherp bau",
        "mense-korte", "albert weil", "gehrken straßen", "schüssler straßen",
        "kiel straßen", "thelen gruppe", "heinrich meier tief",
        "ehlhardt tief", "nettelbeck bau", "haddick",
        # PLZ 1 (Berlin/Brandenburg/Mecklenburg-Vorpommern) - regional repair specialists
        "mbn bau", "jean bratengeier", "schälerbau", "gsb straßenbau",
        "kommunalbau gmbh", "rädel & heidemann", "stb tief", "sti potsdam",
        "oderland straßenbau", "märkische straßenbau", "tiefbau barnim",
        "uckermark straßenbau", "nordpflaster", "heinrich rädel",
        "f & s straßen", "dreibrück", "rostocker straßenbau",
        "nbb neubrandenburger", "mecklenburgische straßenbau",
        "straßen- und tiefbau stralsund", "pommersche tiefbau",
        "rügener straßen", "vorpommersche bau", "arge straßenbau güstrow",
        # PLZ 3 (Hannover/Braunschweig/Kassel/Göttingen/Magdeburg) - regional repair specialists
        "helfrich bau", "raulf straßen", "schütte bau", "hungerland",
        "august brötje", "wilhelm fischer", "fricke schüttpelz",
        "lehne bau", "ritter bau", "gehrke straßen", "hermanns straßen",
        "otto alte-teigeler", "nöhre bau", "konrad bau", "schäfer straßenbau",
        "rinne straßen", "oppermann bau", "bruns straßenbau", "wesemann",
        "steinbrecher bau", "hoch- und tiefbau halberstadt",
        "harz-asphaltmischwerk", "stendaler straßen", "mbu magdeburger",
        "schönebecker straßenbau", "harzer hoch",
        "deutsche asphalt", "rohde verkehrsbau",
    ]

    # --- C-LEADS: Engineering offices, federal agencies (always C) ---
    if "ingenieurbüro" in kategorie:
        return "C"
    if "bund" in kategorie or "autobahn" in kategorie:
        return "C"
    if "autobahn" in name:
        return "C"

    # --- A-LEADS: Kommune with phone+email, or repair-focused private companies ---
    if "kommune" in kategorie:
        # A: Must have phone AND (email OR named contact)
        if has_phone and (has_email or has_contact):
            return "A"
        # B: Has phone OR email but not both
        if has_phone or has_email:
            return "B"
        return "B"

    # Private companies with explicit repair/maintenance focus + contact data
    if "privat" in kategorie:
        is_repair_specialist = has_frost_relevance or any(kw in name for kw in high_relevance_names)

        if "großunternehmen" in kategorie:
            # Large companies: B at best (long sales cycles)
            return "B"

        if is_repair_specialist and has_phone and has_email:
            return "A"
        if is_repair_specialist and (has_phone or has_email):
            return "B"
        if has_phone and has_email:
            return "B"
        if has_phone or has_email:
            return "B"
        return "C"  # No contact data at all

    # --- B-LEADS: State authorities, Landkreise ---
    if "behörde" in kategorie or "land " in kategorie or "stba" in kategorie:
        return "B"
    if "landkreis" in kategorie:
        return "B"

    return "B"


def get_gespraechsaufhaenger(rec):
    """Generate a conversation starter based on category."""
    kategorie = rec["kategorie"].lower()

    if "kommune" in kategorie:
        return "Frostschäden auf Gemeindestraßen schnell beheben - Kaltasphalt ARM als Sofortlösung für Ihren Bauhof"
    elif "privat" in kategorie:
        return "Ergänzung für schnelle Kleinreparaturen zwischen Heißasphalt-Einsätzen - ARM Kaltasphalt"
    elif "behörde" in kategorie or "land" in kategorie or "stba" in kategorie:
        return "Sofort-Reparaturlösung für Straßenunterhalt nach Frostperiode - ARM Asphalt Repair Mix"
    elif "ingenieurbüro" in kategorie:
        return "Innovatives Kaltasphalt-Produkt für Ihre Instandsetzungsprojekte - ARM Asphalt Repair Mix"
    else:
        return "Schnelle Frostschaden-Reparatur mit ARM Asphalt Repair Mix"


# ---------------------------------------------------------------------------
# 4. NAME SPLITTING
# ---------------------------------------------------------------------------

def split_contact_name(kontaktperson):
    """Split contact person into first name and last name."""
    if not kontaktperson or kontaktperson in ["–", "-", ""]:
        return "", ""

    # Remove role in parentheses
    name = re.sub(r'\(.*?\)', '', kontaktperson).strip()
    # Remove titles
    for title in ["Dr. Ing.", "Dr.-Ing.", "Dr.", "Dipl.-Ing.", "Prof.",
                   "Herr", "Frau", "Ing."]:
        name = name.replace(title, "").strip()

    parts = name.split()
    if len(parts) == 0:
        return "", ""
    elif len(parts) == 1:
        return "", parts[0]
    else:
        return parts[0], " ".join(parts[1:])


def extract_role(kontaktperson):
    """Extract role/position from contact person string."""
    if not kontaktperson:
        return ""
    match = re.search(r'\(([^)]+)\)', kontaktperson)
    if match:
        role = match.group(1)
        # Normalize common roles
        role_map = {
            "GF": "Geschäftsführer",
            "Vertretung": "Vertretungsberechtigter",
            "vertretungsberechtigt": "Vertretungsberechtigter",
            "Amtsleitung": "Amtsleiter",
            "Amtsleiter": "Amtsleiter",
            "Fachbereichsleiter": "Fachbereichsleiter",
            "Direktor": "Direktor",
            "Kontakt": "Kontakt",
        }
        return role_map.get(role, role)
    return ""


def determine_abteilung(rec):
    """Determine department based on category."""
    kategorie = rec["kategorie"].lower()
    name = rec["name"].lower()

    if "tiefbau" in name or "tiefbau" in kategorie:
        return "Tiefbau"
    elif "straßenbau" in name or "straßen" in name:
        return "Straßenbau"
    elif "bauhof" in name:
        return "Bauhof"
    elif "ingenieurbüro" in kategorie:
        return "Planung"
    elif "kommune" in kategorie:
        return "Tiefbau"
    elif "behörde" in kategorie:
        return "Straßenbau"
    else:
        return "Straßenbau"


# ---------------------------------------------------------------------------
# 5. EXPORT
# ---------------------------------------------------------------------------

def export_crm_csv(records, filepath):
    """Export leads in Salesforce CRM import format."""
    fieldnames = [
        "Name", "FirstName", "LastName", "PostalCode", "City",
        "Email", "Phone", "MobilePhone", "Company", "Title",
        "IBS_SC_Position__c", "IBS_SC_Abteilungen__c"
    ]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()

        for rec in records:
            first, last = split_contact_name(rec["kontaktperson"])
            if not last:
                last = rec["name"][:40]  # Use company name if no contact person

            writer.writerow({
                "Name": f"{first} {last}".strip() if first else last,
                "FirstName": first,
                "LastName": last,
                "PostalCode": rec["plz"],
                "City": rec["ort"],
                "Email": rec["email"],
                "Phone": rec["telefon"],
                "MobilePhone": "",
                "Company": rec["name"],
                "Title": f"{rec['priority']}-Lead ARM",
                "IBS_SC_Position__c": extract_role(rec["kontaktperson"]),
                "IBS_SC_Abteilungen__c": determine_abteilung(rec),
            })

    print(f"CRM-Import CSV exportiert: {filepath}")


def export_anrufliste_csv(records, filepath):
    """Export prioritized call list."""
    fieldnames = [
        "Priorität", "Firma", "Kategorie", "Ansprechpartner",
        "Telefon", "Email", "PLZ", "Ort", "Gesprächsaufhänger", "Notiz"
    ]

    # Sort: A first, then B, then C
    priority_order = {"A": 0, "B": 1, "C": 2}
    sorted_records = sorted(records, key=lambda r: (
        priority_order.get(r["priority"], 9),
        r["ort"]
    ))

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()

        for rec in sorted_records:
            writer.writerow({
                "Priorität": rec["priority"],
                "Firma": rec["name"],
                "Kategorie": rec["kategorie"],
                "Ansprechpartner": rec["kontaktperson"],
                "Telefon": rec["telefon"],
                "Email": rec["email"],
                "PLZ": rec["plz"],
                "Ort": rec["ort"],
                "Gesprächsaufhänger": get_gespraechsaufhaenger(rec),
                "Notiz": rec.get("leistungen", ""),
            })

    print(f"Anrufliste CSV exportiert: {filepath}")


def export_excel(records, filepath):
    """Export comprehensive Excel with A/B/C tabs."""
    if not HAS_OPENPYXL:
        print("Excel-Export übersprungen (openpyxl nicht installiert)")
        return

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fills = {
        "A": PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),  # dark green
        "B": PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),  # dark orange
        "C": PatternFill(start_color="37474F", end_color="37474F", fill_type="solid"),  # dark grey
        "Übersicht": PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),  # blue
    }
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    columns = [
        ("Firma", 40), ("Kategorie", 25), ("Ansprechpartner", 25),
        ("Position", 20), ("Telefon", 22), ("Email", 35),
        ("Straße", 30), ("PLZ", 8), ("Ort", 18), ("Bundesland", 15),
        ("Leistungen", 50), ("Gesprächsaufhänger", 55),
    ]

    priority_order = {"A": 0, "B": 1, "C": 2}
    sorted_records = sorted(records, key=lambda r: (
        priority_order.get(r["priority"], 9),
        r["ort"]
    ))

    # Create overview sheet
    ws_overview = wb.active
    ws_overview.title = "Übersicht"
    ws_overview.append(["ARM Kampagne - Leadliste", "", "", f"Stand: Februar 2026"])
    ws_overview.append([])
    ws_overview.append(["Priorität", "Anzahl", "Beschreibung"])

    a_count = sum(1 for r in records if r["priority"] == "A")
    b_count = sum(1 for r in records if r["priority"] == "B")
    c_count = sum(1 for r in records if r["priority"] == "C")

    ws_overview.append(["A-Leads", a_count,
                         "Kommunale Tiefbauämter + kleine/mittlere Straßenbauer mit Sanierungsfokus"])
    ws_overview.append(["B-Leads", b_count,
                         "Größere Bauunternehmen + Staatliche Bauämter"])
    ws_overview.append(["C-Leads", c_count,
                         "Ingenieurbüros + Autobahn GmbH + ohne Kontaktdaten"])
    ws_overview.append(["GESAMT", a_count + b_count + c_count, ""])
    ws_overview.append([])
    ws_overview.append(["Angebot: Kaufe 2 Paletten (2x48 Sack), davon 24 Sack ohne Berechnung!"])
    ws_overview.append(["Laufzeit: Ende März (Option Ende April)"])

    # Style overview
    for cell in ws_overview[1]:
        if cell.value:
            cell.font = Font(bold=True, size=14)
    for cell in ws_overview[3]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = header_fills["Übersicht"]
            cell.font = header_font
    ws_overview.column_dimensions['A'].width = 15
    ws_overview.column_dimensions['B'].width = 10
    ws_overview.column_dimensions['C'].width = 70

    # Create A/B/C tabs
    for priority in ["A", "B", "C"]:
        priority_records = [r for r in sorted_records if r["priority"] == priority]
        label = {"A": "A-Leads (Sofort)", "B": "B-Leads (Zweite Welle)", "C": "C-Leads (Optional)"}
        ws = wb.create_sheet(title=label[priority])

        # Header row
        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fills[priority]
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width

        # Data rows
        for row_idx, rec in enumerate(priority_records, 2):
            values = [
                rec["name"],
                rec["kategorie"],
                rec["kontaktperson"],
                extract_role(rec["kontaktperson"]),
                rec["telefon"],
                rec["email"],
                rec["strasse"],
                rec["plz"],
                rec["ort"],
                rec["bundesland"],
                rec["leistungen"],
                get_gespraechsaufhaenger(rec),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Freeze header row
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}1"

    wb.save(filepath)
    print(f"Excel exportiert: {filepath}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("ARM Kampagne - Leadliste Generator (Multi-Region)")
    print("=" * 60)

    # Step 1: Extract data from all configured PLZ regions
    print("\n[1/5] Daten extrahieren...")
    all_records = []

    for region_key, region_cfg in PLZ_REGIONS.items():
        print(f"\n  --- {region_cfg['label']} ---")

        # ChatGPT report (if available)
        if region_cfg["chatgpt"]:
            chatgpt_file = SCRIPT_DIR / region_cfg["chatgpt"]
            if chatgpt_file.exists():
                chatgpt_records = parse_chatgpt_csv(chatgpt_file)
                print(f"  ChatGPT-Report: {len(chatgpt_records)} Einträge")
                all_records.extend(chatgpt_records)
            else:
                print(f"  ChatGPT-Report: Datei nicht gefunden ({chatgpt_file.name})")

        # Claude report
        if region_cfg["claude"]:
            claude_file = SCRIPT_DIR / region_cfg["claude"]
            if claude_file.exists():
                claude_records = parse_claude_markdown_tables(claude_file)
                print(f"  Claude-Report:  {len(claude_records)} Einträge")
                all_records.extend(claude_records)
            else:
                print(f"  Claude-Report:  Datei nicht gefunden ({claude_file.name})")

    print(f"\n  Gesamt (roh):   {len(all_records)} Einträge aus {len(PLZ_REGIONS)} Regionen")

    # Step 1b: Clean up data
    for rec in all_records:
        # Fix approximate PLZ like "90xxx" -> leave empty (better than wrong)
        if rec["plz"] and "x" in rec["plz"].lower():
            rec["plz"] = ""
        # Clean city names: remove PLZ prefix from "90xxx Nürnberg" style
        rec["ort"] = re.sub(r'\*+', '', rec["ort"]).strip()
        rec["ort"] = re.sub(r'^\d+[x]*\s+', '', rec["ort"]).strip()
        # Clean email: remove non-email values like "(Kontaktformular)"
        if rec["email"] and "@" not in rec["email"]:
            rec["email"] = ""
        # Clean kontaktperson: remove non-person values
        if rec["kontaktperson"]:
            non_person = ["(keine person", "(kein ap)", "sekretariat"]
            if any(np in rec["kontaktperson"].lower() for np in non_person):
                rec["kontaktperson"] = ""
        # Auto-detect Bundesland from PLZ if not set
        if not rec["bundesland"] and rec["plz"]:
            rec["bundesland"] = detect_bundesland(rec["plz"])

    # Step 2: Deduplicate
    print("\n[2/5] Deduplizieren...")
    unique_records = deduplicate(all_records)
    print(f"  Nach Dedup:     {len(unique_records)} Einträge")
    print(f"  Entfernt:       {len(all_records) - len(unique_records)} Duplikate")

    # Step 3: Prioritize
    print("\n[3/5] Priorisierung (A/B/C)...")
    for rec in unique_records:
        rec["priority"] = categorize_priority(rec)

    a_leads = [r for r in unique_records if r["priority"] == "A"]
    b_leads = [r for r in unique_records if r["priority"] == "B"]
    c_leads = [r for r in unique_records if r["priority"] == "C"]

    print(f"  A-Leads (Sofort anrufen):    {len(a_leads)}")
    print(f"  B-Leads (Zweite Welle):      {len(b_leads)}")
    print(f"  C-Leads (Optional):          {len(c_leads)}")

    # Breakdown by Bundesland
    bundeslaender = {}
    for rec in unique_records:
        bl = rec.get("bundesland", "Unbekannt") or "Unbekannt"
        bundeslaender[bl] = bundeslaender.get(bl, 0) + 1
    print("\n  Verteilung nach Bundesland:")
    for bl, count in sorted(bundeslaender.items()):
        print(f"    {bl:<25} {count}")

    # Step 4: Export CRM CSV
    print("\n[4/5] CRM-Import CSV exportieren...")
    crm_path = OUTPUT_DIR / "ARM_CRM_Import_Leads.csv"
    export_crm_csv(unique_records, crm_path)

    # Step 5: Export Anrufliste
    print("\n[5/5] Anrufliste + Excel exportieren...")
    call_path = OUTPUT_DIR / "ARM_Anrufliste_Priorisiert.csv"
    export_anrufliste_csv(unique_records, call_path)

    excel_path = OUTPUT_DIR / "ARM_Leadliste_Komplett.xlsx"
    export_excel(unique_records, excel_path)

    # Summary
    print("\n" + "=" * 60)
    print("FERTIG!")
    print("=" * 60)
    print(f"\nDateien:")
    print(f"  1. {crm_path.name:<40} (Salesforce CRM Import)")
    print(f"  2. {call_path.name:<40} (Priorisierte Anrufliste)")
    if HAS_OPENPYXL:
        print(f"  3. {excel_path.name:<40} (Excel mit A/B/C Tabs)")

    print(f"\n--- Top 10 A-Leads ---")
    for i, rec in enumerate(sorted(a_leads, key=lambda r: r["ort"]), 1):
        contact = rec["kontaktperson"][:30] if rec["kontaktperson"] else "(kein AP)"
        print(f"  {i:2d}. {rec['name'][:45]:<45} | {rec['ort']:<15} | {contact}")
        if i >= 10:
            break


if __name__ == "__main__":
    main()
